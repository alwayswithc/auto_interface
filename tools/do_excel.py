from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools.project_path import *
class DoExcel:
    @staticmethod
    def get_data(file_name):
        wb=load_workbook(file_name)
        mode=eval(ReadConfig.read_config(test_config_path,'MODE','mode'))
        test_data=[]
        for key in mode: #遍历配置文件中的字典（键）
            sheet=wb[key]
            if mode[key]=='all':
                for i in range(2,sheet.max_row+1):
                    row_data={}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['method'] = sheet.cell(i, 2).value
                    row_data['url'] = sheet.cell(i, 3).value
                    row_data['data'] = sheet.cell(i, 4).value
                    row_data['expected'] = sheet.cell(i, 5).value
                    row_data['sheet_name']=sheet.cell(i,6).value
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['method'] = sheet.cell(case_id+1, 2).value
                    row_data['url'] = sheet.cell(case_id+1, 3).value
                    row_data['data'] = sheet.cell(case_id+1, 4).value
                    row_data['expected'] = sheet.cell(case_id+1, 5).value
                    row_data['sheet_name'] = sheet.cell(case_id+1, 6).value
                    test_data.append(row_data)
            return test_data

    @staticmethod
    def write_back(file_name,sheet_name,i,result,test_result):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(i,7).value=result
        sheet.cell(i, 8).value =test_result
        wb.save(file_name)
if __name__ == '__main__':
    print(DoExcel.get_data(test_case_path))



